"""
etable5_code.py -- Dental follow-up change analyses and landmark Cox models.

Run 00_setup_and_mice.py first. The Cox models in Panel C are fitted
separately in all 20 imputed datasets and pooled with Rubin's rules.

Outputs:
  etable5_output.xlsx
  etable5_panelA.csv
  etable5_panelB.csv
  etable5_panelC.csv
  etable5_mi_audit.csv
"""

from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats

from coxph import fit_coxph
from pooling import rubin_pool_vector


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
DATA_DIR = Path(".")
FOLLOWUP_DETAIL = DATA_DIR / "KFACS_추적조사상세_최종_260608__2_.xlsx"

MI_FILE = DATA_DIR / "mi_long_v2.pkl"
DENTAL_W2_FILE = DATA_DIR / "dental_followup_2018_2019.pkl"
DENTAL_W3_FILE = DATA_DIR / "dental_wave2020.pkl"

REFERENCE_IMPUTATION = 1

required = [MI_FILE, DENTAL_W2_FILE, DENTAL_W3_FILE, FOLLOWUP_DETAIL]
missing = [str(p) for p in required if not p.exists()]
if missing:
    raise FileNotFoundError(
        "Required files were not found. Run 00_setup_and_mice.py first and "
        "place the follow-up detail workbook in DATA_DIR. Missing: "
        + ", ".join(missing)
    )


# ---------------------------------------------------------------------------
# Load imputed analytic data
# ---------------------------------------------------------------------------
mi = pd.read_pickle(MI_FILE).copy()

if "_imputation_id" not in mi.columns:
    raise ValueError("mi_long_v2.pkl does not contain _imputation_id.")

imputation_ids = sorted(mi["_imputation_id"].dropna().astype(int).unique())
if len(imputation_ids) != 20:
    print(f"Warning: expected 20 imputations, found {len(imputation_ids)}.")


# ---------------------------------------------------------------------------
# Landmark-time scaffolds and observed follow-up dental data
# ---------------------------------------------------------------------------
def numeric_frame(df, columns):
    out = df.copy()
    for c in columns:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
    return out


def build_landmark_age(sheet_name, age_column):
    bl = pd.read_excel(FOLLOWUP_DETAIL, sheet_name="BL").rename(columns={"No.": "id"})
    fu = pd.read_excel(FOLLOWUP_DETAIL, sheet_name=sheet_name).rename(columns={"No.": "id"})

    keep_bl = ["id", "BL_age_365"]
    keep_fu = ["id", age_column]

    bl = numeric_frame(bl[keep_bl], ["BL_age_365"])
    fu = numeric_frame(fu[keep_fu], [age_column])

    out = bl.merge(fu, on="id", how="inner")
    out["landmark_time"] = out[age_column] - out["BL_age_365"]
    out = out.dropna(subset=["landmark_time"])
    out = out[out["landmark_time"] >= 0].copy()
    return out[["id", "landmark_time"]]


w2_age = build_landmark_age("W2(2018-2019)", "W2_age_365")

dent_w2 = pd.read_pickle(DENTAL_W2_FILE).copy()
dent_w2 = dent_w2.rename(
    columns={
        "natural_teeth": "natural_teeth_w2",
        "dental_implant_count": "implant_w2",
    }
)
dent_w2 = numeric_frame(dent_w2, ["natural_teeth_w2", "implant_w2"])
dent_w2.loc[~dent_w2["natural_teeth_w2"].between(0, 32), "natural_teeth_w2"] = np.nan
dent_w2.loc[dent_w2["implant_w2"] < 0, "implant_w2"] = np.nan
dent_w2["fd_w2"] = (
    dent_w2["natural_teeth_w2"].fillna(0)
    + dent_w2["implant_w2"].fillna(0)
).clip(upper=32)
dent_w2.loc[dent_w2["natural_teeth_w2"].isna(), "fd_w2"] = np.nan

w2_scaffold = (
    w2_age.merge(
        dent_w2[["id", "natural_teeth_w2", "implant_w2", "fd_w2"]],
        on="id",
        how="inner",
    )
    .dropna(subset=["natural_teeth_w2", "fd_w2"])
    .drop_duplicates("id")
)


# Later dental follow-up subset used in Panel A
w3_age = build_landmark_age("W3(2020-2021)", "W3_age_365")

dent_w3 = pd.read_pickle(DENTAL_W3_FILE).copy()
dent_w3["natural_teeth_w3"] = pd.to_numeric(
    dent_w3["natural_teeth"], errors="coerce"
)
dent_w3.loc[
    ~dent_w3["natural_teeth_w3"].between(0, 32), "natural_teeth_w3"
] = np.nan
dent_w3 = (
    dent_w3.loc[dent_w3["natural_teeth_w3"].notna(), ["id", "natural_teeth_w3"]]
    .drop_duplicates("id")
)

w3_scaffold = (
    w3_age.merge(dent_w3, on="id", how="inner")
    .dropna(subset=["natural_teeth_w3"])
    .drop_duplicates("id")
)


# ---------------------------------------------------------------------------
# Construct Wave-2 landmark cohort separately in each imputed dataset
# ---------------------------------------------------------------------------
CENTER_DUMMIES = [f"center_{i}" for i in range(2, 10)]

MODEL1 = ["age", "female", "cohort2017"] + CENTER_DUMMIES

MODEL2 = MODEL1 + [
    "edu",
    "income_middle",
    "income_low",
    "smoking_history_ge100",
    "ever_alcohol_use",
    "systemic_disease_count",
    "systemic_reserve_adverse_count",
    "natural_teeth",
    "dental_implant_count",
]

MODEL3 = MODEL2 + ["natural_teeth_change"]

REDUCED_DEMOGRAPHIC = MODEL1

REDUCED_CLINICAL = MODEL1 + [
    "systemic_disease_count",
    "systemic_reserve_adverse_count",
]


def build_covariates(d):
    d = d.copy()

    d["female"] = (pd.to_numeric(d["sex"], errors="coerce") == 2).astype(float)
    d["cohort2017"] = (
        pd.to_numeric(d["cohort_start_year"], errors="coerce") == 2017
    ).astype(float)

    income = d["income_level"].astype("object")
    d["income_middle"] = (income == "Middle").astype(float)
    d["income_low"] = (income == "Low/no income").astype(float)

    center_num = pd.to_numeric(d["center"], errors="coerce")
    for i in range(2, 10):
        d[f"center_{i}"] = (center_num == i).astype(float)

    d["natural_teeth_change"] = (
        pd.to_numeric(d["natural_teeth_w2"], errors="coerce")
        - pd.to_numeric(d["natural_teeth"], errors="coerce")
    )
    d["implant_gain"] = (
        pd.to_numeric(d["implant_w2"], errors="coerce")
        - pd.to_numeric(d["dental_implant_count"], errors="coerce")
    )
    d["implant_gain_per2"] = d["implant_gain"] / 2.0

    return d


def make_w2_landmark(imputation_id):
    d = mi.loc[mi["_imputation_id"] == imputation_id].copy()
    d = d.merge(w2_scaffold, on="id", how="inner")

    d["followup_years"] = pd.to_numeric(d["followup_years"], errors="coerce")
    d["death_event"] = pd.to_numeric(d["death_event"], errors="coerce")

    # Participants must be alive and under follow-up at the Wave-2 landmark.
    d = d.dropna(subset=["followup_years", "death_event", "landmark_time"])
    d = d[d["followup_years"] >= d["landmark_time"]].copy()

    d["post_landmark_followup"] = d["followup_years"] - d["landmark_time"]
    d["post_landmark_event"] = d["death_event"]

    d["implant_increase"] = (
        pd.to_numeric(d["implant_w2"], errors="coerce")
        > pd.to_numeric(d["dental_implant_count"], errors="coerce")
    ).astype(int)

    d["fd_increase"] = (
        pd.to_numeric(d["fd_w2"], errors="coerce")
        > pd.to_numeric(d["functional_dentition"], errors="coerce")
    ).astype(int)

    d["fd_decrease"] = (
        pd.to_numeric(d["fd_w2"], errors="coerce")
        < pd.to_numeric(d["functional_dentition"], errors="coerce")
    ).astype(int)

    d["new_implant_and_fd_gain"] = (
        (d["implant_increase"] == 1) & (d["fd_increase"] == 1)
    ).astype(int)

    d["baseline_fd_lt20"] = (
        pd.to_numeric(d["functional_dentition"], errors="coerce") < 20
    ).astype(int)

    d["restored_to_fd20"] = (
        (d["baseline_fd_lt20"] == 1)
        & (pd.to_numeric(d["fd_w2"], errors="coerce") >= 20)
    ).astype(int)

    return build_covariates(d)


landmark_by_imp = {
    imp: make_w2_landmark(imp)
    for imp in imputation_ids
}


# ---------------------------------------------------------------------------
# Cox fitting and Rubin pooling
# ---------------------------------------------------------------------------
def fit_single(d, exposure_col, covars, subset=None):
    x = d.copy()

    if subset is not None:
        x = x.loc[subset(x)].copy()

    cols = [exposure_col] + covars
    needed = ["post_landmark_followup", "post_landmark_event"] + cols
    x = x[needed].replace([np.inf, -np.inf], np.nan).dropna()

    if x.empty:
        raise ValueError(f"No complete observations for exposure {exposure_col}.")

    X = x[cols].astype(float).to_numpy()
    time = x["post_landmark_followup"].astype(float).to_numpy()
    event = x["post_landmark_event"].astype(float).to_numpy()

    if np.nanstd(X[:, 0]) == 0:
        raise ValueError(f"Exposure {exposure_col} has no variation in the analysis set.")

    fit = fit_coxph(time, event, X)

    exposed_n = np.nan
    exposed_events = np.nan
    vals = x[exposure_col]

    if set(pd.unique(vals.dropna())).issubset({0, 1, 0.0, 1.0}):
        exposed = vals == 1
        exposed_n = int(exposed.sum())
        exposed_events = int(x.loc[exposed, "post_landmark_event"].sum())

    return {
        "beta": float(fit["beta"][0]),
        "variance": float(fit["cov"][0, 0]),
        "n": int(fit["n"]),
        "events": int(fit["n_events"]),
        "exposed_n": exposed_n,
        "exposed_events": exposed_events,
    }


def rubin_pool_scalar(estimates):
    betas = np.array([[r["beta"]] for r in estimates], dtype=float)
    variances = np.array([[r["variance"]] for r in estimates], dtype=float)

    qbar, se, df = rubin_pool_vector(betas, variances)
    beta = float(qbar[0])
    se = float(se[0])
    df = float(df[0])

    tcrit = stats.t.ppf(0.975, df)
    lo_beta = beta - tcrit * se
    hi_beta = beta + tcrit * se

    tstat = beta / se
    p = 2 * (1 - stats.t.cdf(abs(tstat), df))

    return {
        "beta": beta,
        "se": se,
        "df": df,
        "hr": float(np.exp(beta)),
        "lo": float(np.exp(lo_beta)),
        "hi": float(np.exp(hi_beta)),
        "p": float(p),
    }


def pooled_landmark(exposure_col, covars, subset=None):
    per_imp = []
    for imp in imputation_ids:
        r = fit_single(
            landmark_by_imp[imp],
            exposure_col,
            covars,
            subset=subset,
        )
        r["imputation"] = imp
        per_imp.append(r)

    pooled = rubin_pool_scalar(per_imp)
    return pooled, pd.DataFrame(per_imp)


def subset_implant_increasers(d):
    return d["implant_increase"] == 1


def subset_baseline_fd_lt20(d):
    return d["baseline_fd_lt20"] == 1


# ---------------------------------------------------------------------------
# Panel A: cohort construction
# ---------------------------------------------------------------------------
ref = landmark_by_imp[REFERENCE_IMPUTATION]

ref_full = mi.loc[mi["_imputation_id"] == REFERENCE_IMPUTATION]
baseline_n = int(len(ref_full))
baseline_deaths = int(ref_full["death_event"].sum())

wave2_n = int(len(ref))
wave2_deaths = int(ref["post_landmark_event"].sum())


def make_w3_subset(imputation_id):
    d = mi.loc[mi["_imputation_id"] == imputation_id].copy()
    d = d.merge(w3_scaffold, on="id", how="inner")
    d["followup_years"] = pd.to_numeric(d["followup_years"], errors="coerce")
    d["death_event"] = pd.to_numeric(d["death_event"], errors="coerce")
    d = d.dropna(subset=["followup_years", "death_event", "landmark_time"])
    d = d[d["followup_years"] >= d["landmark_time"]].copy()
    d["post_landmark_event"] = d["death_event"]
    return d


w3_by_imp = {
    imp: make_w3_subset(imp)
    for imp in imputation_ids
}
w3_ref = w3_by_imp[REFERENCE_IMPUTATION]

panel_a = pd.DataFrame(
    [
        {
            "Analysis set": "Main baseline survival cohort",
            "Dental exposure/change period": "Baseline dental assessment, 2016–2017",
            "Mortality ascertainment period": "Baseline to death/censoring during serial KFACS follow-up",
            "Eligible participants, n": baseline_n,
            "Deaths included in analysis, n": baseline_deaths,
            "Analytic role": "Primary analysis",
        },
        {
            "Analysis set": "Wave 2 landmark cohort",
            "Dental exposure/change period": "Wave 1 → Wave 2",
            "Mortality ascertainment period": "Wave 2 landmark to death/censoring during subsequent KFACS follow-up",
            "Eligible participants, n": wave2_n,
            "Deaths included in analysis, n": wave2_deaths,
            "Analytic role": "Primary exploratory landmark analysis",
        },
        {
            "Analysis set": "Later dental follow-up subset",
            "Dental exposure/change period": "Wave 1 → later dental data, 2020",
            "Mortality ascertainment period": "Later landmark to death/censoring during subsequent KFACS follow-up",
            "Eligible participants, n": int(len(w3_ref)),
            "Deaths included in analysis, n": int(w3_ref["post_landmark_event"].sum()),
            "Analytic role": "Sensitivity/feasibility only",
        },
    ]
)


# ---------------------------------------------------------------------------
# Panel B: change phenotype frequencies
# Descriptive counts use the same reference imputation throughout. Variation
# across imputations is reported separately in the MI Audit sheet.
# ---------------------------------------------------------------------------
def count_binary(d, col):
    exposed = d[col] == 1
    n = int(exposed.sum())
    deaths = int(d.loc[exposed, "post_landmark_event"].sum())
    return n, len(d), deaths


panel_b_rows = []

n, N, deaths = count_binary(ref, "implant_increase")
panel_b_rows.append(
    {
        "Change phenotype": "Implant-supported restoration increase",
        "Definition": "Wave 2 implant count > baseline implant count",
        "n/N (%)": f"{n}/{N} ({100*n/N:.1f})",
        "Post-landmark deaths, n/N (%)": f"{deaths}/{n} ({100*deaths/n:.1f})",
    }
)

no_inc = ref["implant_increase"] == 0
n_no = int(no_inc.sum())
d_no = int(ref.loc[no_inc, "post_landmark_event"].sum())
panel_b_rows.append(
    {
        "Change phenotype": "No implant-supported restoration increase",
        "Definition": "Wave 2 implant count ≤ baseline implant count",
        "n/N (%)": f"{n_no}/{len(ref)} ({100*n_no/len(ref):.1f})",
        "Post-landmark deaths, n/N (%)": f"{d_no}/{n_no} ({100*d_no/n_no:.1f})",
    }
)

for col, label, definition in [
    (
        "fd_increase",
        "Functional dentition increase",
        "Wave 2 functional dentition > baseline functional dentition",
    ),
    (
        "new_implant_and_fd_gain",
        "New implant + functional dentition gain",
        "Implant count increased and functional dentition increased",
    ),
]:
    n, N, deaths = count_binary(ref, col)
    panel_b_rows.append(
        {
            "Change phenotype": label,
            "Definition": definition,
            "n/N (%)": f"{n}/{N} ({100*n/N:.1f})",
            "Post-landmark deaths, n/N (%)": f"{deaths}/{n} ({100*deaths/n:.1f})",
        }
    )

ref_lt20 = ref.loc[ref["baseline_fd_lt20"] == 1].copy()
n, N, deaths = count_binary(ref_lt20, "restored_to_fd20")
panel_b_rows.append(
    {
        "Change phenotype": "Restoration to functional dentition ≥20",
        "Definition": "Baseline functional dentition <20 and Wave 2 functional dentition ≥20",
        "n/N (%)": f"{n}/{N} ({100*n/N:.1f})",
        "Post-landmark deaths, n/N (%)": f"{deaths}/{n} ({100*deaths/n:.1f})",
    }
)

n, N, deaths = count_binary(ref, "fd_decrease")
panel_b_rows.append(
    {
        "Change phenotype": "Functional dentition decreases",
        "Definition": "Wave 2 functional dentition < baseline functional dentition",
        "n/N (%)": f"{n}/{N} ({100*n/N:.1f})",
        "Post-landmark deaths, n/N (%)": f"{deaths}/{n} ({100*deaths/n:.1f})",
    }
)

panel_b = pd.DataFrame(panel_b_rows)


# ---------------------------------------------------------------------------
# Panel C: pooled landmark Cox models
# ---------------------------------------------------------------------------
def format_p(p):
    return "<.001" if p < 0.001 else f"{p:.3f}".lstrip("0")


def format_hr(r):
    return f"{r['hr']:.2f} ({r['lo']:.2f}–{r['hi']:.2f})"


panel_c_specs = [
    {
        "Exposure/change phenotype": "Implant-supported restoration increase",
        "Comparison / estimand": "Increase vs no increase",
        "Model": "Model 1: demographic",
        "exposure": "implant_increase",
        "covars": MODEL1,
        "subset": None,
    },
    {
        "Exposure/change phenotype": "Implant-supported restoration increase",
        "Comparison / estimand": "Increase vs no increase",
        "Model": "Model 2: socioeconomic/clinical + baseline oral",
        "exposure": "implant_increase",
        "covars": MODEL2,
        "subset": None,
    },
    {
        "Exposure/change phenotype": "Implant-supported restoration increase",
        "Comparison / estimand": "Increase vs no increase",
        "Model": "Model 3: Model 2 + natural tooth change",
        "exposure": "implant_increase",
        "covars": MODEL3,
        "subset": None,
    },
    {
        "Exposure/change phenotype": "Positive implant gain",
        "Comparison / estimand": "Per 2 additional implant-supported restorations among those with increase",
        "Model": "Model 2",
        "exposure": "implant_gain_per2",
        "covars": MODEL2,
        "subset": subset_implant_increasers,
    },
    {
        "Exposure/change phenotype": "Positive implant gain",
        "Comparison / estimand": "Per 2 additional implant-supported restorations among those with increase",
        "Model": "Model 3",
        "exposure": "implant_gain_per2",
        "covars": MODEL3,
        "subset": subset_implant_increasers,
    },
    {
        "Exposure/change phenotype": "Functional dentition increase",
        "Comparison / estimand": "Increase vs no increase",
        "Model": "Model 2",
        "exposure": "fd_increase",
        "covars": MODEL2,
        "subset": None,
    },
    {
        "Exposure/change phenotype": "New implant + functional dentition gain",
        "Comparison / estimand": "Yes vs no",
        "Model": "Model 2",
        "exposure": "new_implant_and_fd_gain",
        "covars": MODEL2,
        "subset": None,
    },
    {
        "Exposure/change phenotype": "Restoration to functional dentition ≥20",
        "Comparison / estimand": "Baseline FD <20 and Wave 2 FD ≥20 vs all others",
        "Model": "Model 2",
        "exposure": "restored_to_fd20",
        "covars": MODEL2,
        "subset": None,
    },
    {
        "Exposure/change phenotype": "Restoration to functional dentition ≥20 among baseline FD <20",
        "Comparison / estimand": "Restored to ≥20 vs remained <20",
        "Model": "Reduced demographic model",
        "exposure": "restored_to_fd20",
        "covars": REDUCED_DEMOGRAPHIC,
        "subset": subset_baseline_fd_lt20,
    },
    {
        "Exposure/change phenotype": "Restoration to functional dentition ≥20 among baseline FD <20",
        "Comparison / estimand": "Restored to ≥20 vs remained <20",
        "Model": "Reduced clinical model",
        "exposure": "restored_to_fd20",
        "covars": REDUCED_CLINICAL,
        "subset": subset_baseline_fd_lt20,
    },
]

panel_c_rows = []
audit_parts = []

for spec in panel_c_specs:
    pooled, per_imp = pooled_landmark(
        spec["exposure"],
        spec["covars"],
        subset=spec["subset"],
    )

    per_imp = per_imp.copy()
    per_imp["Analysis"] = (
        spec["Exposure/change phenotype"]
        + " | "
        + spec["Model"]
    )
    audit_parts.append(per_imp)

    ref_row = per_imp.loc[
        per_imp["imputation"] == REFERENCE_IMPUTATION
    ].iloc[0]

    exposed_text = "—"
    if not pd.isna(ref_row["exposed_n"]):
        exposed_text = (
            f"{int(ref_row['exposed_n'])}/{int(ref_row['exposed_events'])}"
        )

    panel_c_rows.append(
        {
            "Exposure/change phenotype": spec["Exposure/change phenotype"],
            "Comparison / estimand": spec["Comparison / estimand"],
            "Model": spec["Model"],
            "n/events": f"{int(ref_row['n'])}/{int(ref_row['events'])}",
            "Exposed/events": exposed_text,
            "HR (95% CI)": format_hr(pooled),
            "P value": format_p(pooled["p"]),
        }
    )

panel_c = pd.DataFrame(panel_c_rows)


# ---------------------------------------------------------------------------
# Multiple-imputation audit
# ---------------------------------------------------------------------------
model_audit = pd.concat(audit_parts, ignore_index=True)

cohort_audit_rows = []
for imp in imputation_ids:
    d = landmark_by_imp[imp]
    d3 = w3_by_imp[imp]
    cohort_audit_rows.append(
        {
            "imputation": imp,
            "wave2_n": len(d),
            "wave2_events": int(d["post_landmark_event"].sum()),
            "implant_increase_n": int(d["implant_increase"].sum()),
            "fd_increase_n": int(d["fd_increase"].sum()),
            "new_implant_and_fd_gain_n": int(d["new_implant_and_fd_gain"].sum()),
            "baseline_fd_lt20_n": int(d["baseline_fd_lt20"].sum()),
            "restored_to_fd20_n": int(
                d.loc[d["baseline_fd_lt20"] == 1, "restored_to_fd20"].sum()
            ),
            "fd_decrease_n": int(d["fd_decrease"].sum()),
            "wave3_n": len(d3),
            "wave3_events": int(d3["post_landmark_event"].sum()),
        }
    )

cohort_audit = pd.DataFrame(cohort_audit_rows)

mi_audit = model_audit.merge(
    cohort_audit,
    on="imputation",
    how="left",
)


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
panel_a.to_csv(DATA_DIR / "etable5_panelA.csv", index=False, encoding="utf-8-sig")
panel_b.to_csv(DATA_DIR / "etable5_panelB.csv", index=False, encoding="utf-8-sig")
panel_c.to_csv(DATA_DIR / "etable5_panelC.csv", index=False, encoding="utf-8-sig")
mi_audit.to_csv(DATA_DIR / "etable5_mi_audit.csv", index=False, encoding="utf-8-sig")

xlsx_path = DATA_DIR / "etable5_output.xlsx"

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    panel_a.to_excel(writer, sheet_name="Panel A", index=False)
    panel_b.to_excel(writer, sheet_name="Panel B", index=False)
    panel_c.to_excel(writer, sheet_name="Panel C", index=False)
    mi_audit.to_excel(writer, sheet_name="MI Audit", index=False)

    from openpyxl.styles import Alignment, Font

    for sheet in ["Panel A", "Panel B", "Panel C", "MI Audit"]:
        ws = writer.book[sheet]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    widths = {
        "Panel A": [34, 40, 62, 24, 26, 34],
        "Panel B": [48, 72, 22, 32],
        "Panel C": [56, 68, 48, 16, 18, 24, 14],
        "MI Audit": [14, 15, 15, 15, 15, 70, 16, 16, 22, 18, 28, 20, 15, 15],
    }

    for sheet, values in widths.items():
        ws = writer.book[sheet]
        for idx, width in enumerate(values, start=1):
            ws.column_dimensions[
                __import__("openpyxl").utils.get_column_letter(idx)
            ].width = width


print("eTable 5 completed.")
print(f"Imputations pooled: {len(imputation_ids)}")
print(
    "Wave-2 landmark cohort across imputations: "
    f"n {cohort_audit['wave2_n'].min()}–{cohort_audit['wave2_n'].max()}, "
    f"events {cohort_audit['wave2_events'].min()}–{cohort_audit['wave2_events'].max()}"
)
print(f"Wrote: {xlsx_path}")
print("Wrote: etable5_panelA.csv")
print("Wrote: etable5_panelB.csv")
print("Wrote: etable5_panelC.csv")
print("Wrote: etable5_mi_audit.csv")
